"""文档处理服务"""
import os
import uuid
import re
from typing import List, Dict, Any, Optional
import jieba
import pypdf
import docx
import openpyxl
from config import Config

class DocumentProcessor:
    """文档处理器"""
    
    def __init__(self):
        self.chunk_size = Config.CHUNK_SIZE
        self.chunk_overlap = Config.CHUNK_OVERLAP
    
    def extract_text_from_file(self, file_path: str, file_type: str) -> str:
        """从文件提取文本内容"""
        try:
            if file_type.lower() == 'txt':
                return self._extract_from_txt(file_path)
            elif file_type.lower() == 'pdf':
                return self._extract_from_pdf(file_path)
            elif file_type.lower() == 'docx':
                return self._extract_from_docx(file_path)
            elif file_type.lower() == 'xlsx':
                return self._extract_from_xlsx(file_path)
            elif file_type.lower() == 'md':
                return self._extract_from_txt(file_path)  # Markdown作为文本处理
            else:
                raise ValueError(f"不支持的文件类型: {file_type}")
        except Exception as e:
            raise Exception(f"文件内容提取失败: {str(e)}")
    
    def _extract_from_txt(self, file_path: str) -> str:
        """从TXT文件提取文本"""
        encodings = ['utf-8', 'gbk', 'gb2312', 'utf-16']
        for encoding in encodings:
            try:
                with open(file_path, 'r', encoding=encoding) as file:
                    return file.read()
            except UnicodeDecodeError:
                continue
        raise Exception("无法识别文件编码")
    
    def _extract_from_pdf(self, file_path: str) -> str:
        """从PDF文件提取文本"""
        text = ""
        with open(file_path, 'rb') as file:
            pdf_reader = pypdf.PdfReader(file)
            for page in pdf_reader.pages:
                text += page.extract_text() + "\n"
        return text
    
    def _extract_from_docx(self, file_path: str) -> str:
        """从DOCX文件提取文本"""
        doc = docx.Document(file_path)
        text = ""
        for paragraph in doc.paragraphs:
            text += paragraph.text + "\n"
        return text
    
    def _extract_from_xlsx(self, file_path: str) -> str:
        """从XLSX文件提取文本"""
        workbook = openpyxl.load_workbook(file_path)
        text = ""
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text += f"工作表: {sheet_name}\n"
            
            for row in sheet.iter_rows():
                row_text = []
                for cell in row:
                    if cell.value is not None:
                        row_text.append(str(cell.value))
                if row_text:
                    text += "\t".join(row_text) + "\n"
            text += "\n"
        
        return text
    
    def split_text_into_chunks(self, text: str) -> List[Dict[str, Any]]:
        """将文本分割成块"""
        if not text.strip():
            return []
        
        # 清理文本
        text = self._clean_text(text)
        
        # 按句子分割（中文按句号、问号、感叹号等分割）
        sentences = self._split_sentences(text)
        
        chunks = []
        current_chunk = ""
        current_length = 0
        chunk_index = 0
        
        for sentence in sentences:
            sentence_length = len(sentence)
            
            # 如果当前块加上新句子超过块大小，则创建新块
            if current_length + sentence_length > self.chunk_size and current_chunk:
                chunks.append({
                    'chunk_index': chunk_index,
                    'content': current_chunk.strip(),
                    'vector_id': str(uuid.uuid4())
                })
                chunk_index += 1
                
                # 考虑重叠部分
                if self.chunk_overlap > 0:
                    overlap_text = self._get_overlap_text(current_chunk, self.chunk_overlap)
                    current_chunk = overlap_text + sentence
                    current_length = len(current_chunk)
                else:
                    current_chunk = sentence
                    current_length = sentence_length
            else:
                current_chunk += sentence
                current_length += sentence_length
        
        # 添加最后一个块
        if current_chunk.strip():
            chunks.append({
                'chunk_index': chunk_index,
                'content': current_chunk.strip(),
                'vector_id': str(uuid.uuid4())
            })
        
        return chunks
    
    def _clean_text(self, text: str) -> str:
        """清理文本"""
        # 移除多余的空白字符
        text = re.sub(r'\s+', ' ', text)
        # 移除特殊字符但保留中文标点
        text = re.sub(r'[^\w\s\u4e00-\u9fff，。！？；：""''（）【】《》]', '', text)
        return text.strip()
    
    def _split_sentences(self, text: str) -> List[str]:
        """按句子分割文本"""
        # 中英文句子分割符
        sentence_endings = r'[。！？!?;；]\s*'
        sentences = re.split(sentence_endings, text)
        
        # 过滤空句子并保留有意义的内容
        meaningful_sentences = []
        for sentence in sentences:
            sentence = sentence.strip()
            if len(sentence) > 10:  # 只保留长度超过10的句子
                meaningful_sentences.append(sentence + '。')  # 添加句号
        
        # 如果没有明显的句子边界，按长度分割
        if len(meaningful_sentences) < 2:
            return self._split_by_length(text, self.chunk_size // 2)
        
        return meaningful_sentences
    
    def _split_by_length(self, text: str, max_length: int) -> List[str]:
        """按长度分割文本"""
        sentences = []
        for i in range(0, len(text), max_length):
            sentences.append(text[i:i + max_length])
        return sentences
    
    def _get_overlap_text(self, text: str, overlap_size: int) -> str:
        """获取重叠文本"""
        if len(text) <= overlap_size:
            return text
        return text[-overlap_size:]
    
    def validate_file(self, file_path: str, file_type: str) -> Dict[str, Any]:
        """验证文件"""
        result = {
            'valid': True,
            'error': None,
            'file_size': 0
        }
        
        try:
            if not os.path.exists(file_path):
                result['valid'] = False
                result['error'] = "文件不存在"
                return result
            
            # 检查文件大小
            file_size = os.path.getsize(file_path)
            result['file_size'] = file_size
            
            if file_size > Config.MAX_FILE_SIZE:
                result['valid'] = False
                result['error'] = f"文件大小超过限制 ({Config.MAX_FILE_SIZE / 1024 / 1024}MB)"
                return result
            
            if file_size == 0:
                result['valid'] = False
                result['error'] = "文件为空"
                return result
            
            # 检查文件类型
            if file_type.lower() not in Config.ALLOWED_EXTENSIONS:
                result['valid'] = False
                result['error'] = f"不支持的文件类型，支持的类型：{', '.join(Config.ALLOWED_EXTENSIONS)}"
                return result
            
            return result
            
        except Exception as e:
            result['valid'] = False
            result['error'] = f"文件验证失败: {str(e)}"
            return result

# 全局文档处理器实例
document_processor = DocumentProcessor()
